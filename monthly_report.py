import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
import pyodbc
import getpass
import yaml
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import numbers
import argparse
from my_db_module import get_connection

# --- CONFIGURATION ---
now = datetime.now()
today = now.date()
yyyymm = today.strftime("%Y%m")

username = os.getlogin()
base_dir = f"C:\\Users\\{username}\\Desktop\\monthly_report"

# --- Parse Command line arguments for config ---
parser = argparse.ArgumentParser(description = "Run monthnly report with a specified YAML config.")
parser.add_argument(
    "--config",
    default=os.path.join(base_dir, "report_config.yaml"),
    help="Path to the YAML configuration file (default: report_config.yaml in base_dir)"
)
parser.add_argument(
    "--env",
    default="prod",
    help="Environment to use (e.g., prod, test, dev)"
)
args=parser.parse_args()

# ---configure paths ---
config_path = args.config
env = args.env.lower()
CHUNK_SIZE = 50000

# --- Clean illegal characters from cell values ---
def clean_cell_value(value):
    if isinstance(value, str):
        # Replace each illegal character with a question mark
        return ILLEGAL_CHARACTERS_RE.sub("?", value)
    return value

# --- Date range logic ---
def get_date_range(rspec, today):
    end_date = today
    if rspec["type"] == "trailing_months":
        months = rspec["months"]
        start_date = (end_date.replace(day=1) - pd.DateOffset(months=months)).date()
    elif rspec["type"] == "fixed":
        start_val = rspec["start"]
        if isinstance(start_val, str):
           start_date = datetime.strptime(start_val, "%Y-%m-%d").date()
        else:
           start_date = start_val
    else:
        raise ValueError("Invalid date_range type")
    return start_date, end_date

# --- Export query result to Excel ---
def export_query_to_xlsx(query, filename, conn, output_dir,template_path,params=None,sheet_name="Sheet1",start_row=6):
    output_path = os.path.join(output_dir, filename)
    print(f"\n[+] Running query and exporting to: {output_path}")

    # --- Load template workbook ---
    if not os.path.isfile(template_path):
        print(f"[!] Template not found: {template_path}")
        return
    
    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
    
    # Convert date strings to date objects
    if params and len(params) >= 2:
        start_date = datetime.strptime(params[0], "%Y-%m-%d").date()
        end_date = datetime.strptime(params[1], "%Y-%m-%d").date()
    else:
        start_date = None
        end_date = datetime.today().date()

    # --- Set fixed placeholders ---
    ws["B2"] = start_date
    ws["B2"].number_format = numbers.FORMAT_DATE_YYYYMMDD2

    ws["D2"] = end_date
    ws["D2"].number_format = numbers.FORMAT_DATE_YYYYMMDD2

    ws["B3"] = end_date 
    ws["B3"].number_format = numbers.FORMAT_DATE_YYYYMMDD2

    row_counter = 0
    for chunk in pd.read_sql(query, conn, params=params, chunksize=CHUNK_SIZE):
        for row in chunk.itertuples(index=False):
            cleaned_row = [clean_cell_value(cell) for cell in row]
            for col_idx, cell_value in enumerate(cleaned_row, start=1):
                ws.cell(row=start_row + row_counter, column=col_idx, value=cell_value)
            row_counter += 1
        print(f"    [✓] Appended chunk of {len(chunk)} rows")

    wb.save(output_path)
    print(f"[✓] Finished: {filename}")

# --- CONNECT TO DATABASE ---
print(f"\n[+] Connecting to environment: {env}")
conn = get_connection(env=env)

def run_reports_from_yaml(conn, config_path):
    with open(config_path, "r") as f:
        config = yaml.safe_load(f)

    # === Load root paths from config
    queries_root = os.path.normpath(config["paths"]["queries_root"])
    templates_root = os.path.normpath(config["paths"]["templates_root"])
    reports_root = os.path.normpath(config["paths"]["reports_root"])

    # === Process each reportTo user ===
    for person in config["reportTo"]:
        name = person["name"]

        for report in person["reports"]:
            topic = report["topic"]
            date_range = report["date_range"]

            # === Build paths ===
            sql_path = os.path.normpath(os.path.join(queries_root, name, report["sql"]))
            template_path = os.path.normpath(os.path.join(templates_root, name, report["template"]))
            output_dir = os.path.normpath(os.path.join(reports_root, name, yyyymm))
            os.makedirs(output_dir, exist_ok=True)
            output_filename = f"{topic}_{yyyymm}_{name}.xlsx"
            output_path = os.path.normpath(os.path.join(output_dir, output_filename))

            print(f"\n=== Processing report: {topic} for {name} ===")
            print(f"SQL path: {sql_path}")
            print(f"Template path: {template_path}")
            print(f"Output path: {output_path}")

            # === Load and parameterize the SQL ===
            if not os.path.isfile(sql_path):
                print(f"  [!] SQL file not found: {sql_path} — skipping")
                continue

            with open(sql_path, "r", encoding="utf-8") as f:
                query = f.read()

            start_date, end_date = get_date_range(date_range, today)
            print(f"Using date range: {start_date} to {end_date}")

            param_count = query.count("?")
            if param_count >= 2:
                params=[start_date.isoformat(), end_date.isoformat()]
            else:
                params = None

            export_query_to_xlsx(
                query=query,
                filename=output_filename,
                conn=conn,
                output_dir=output_dir,
                template_path=template_path,
                params=params)

if __name__ == "__main__":
    run_reports_from_yaml(conn, config_path)
    conn.close()
    print("\n[✓] All XLSX exports completed.")