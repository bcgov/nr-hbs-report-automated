import os
import re
import gc
from datetime import datetime
from openpyxl import load_workbook
import win32com.client as win32

# Get current user's Desktop path
username = os.getlogin()
base_folder = fr"C:\Users\{username}\Desktop\monthly_report"
template_folder = os.path.join(base_folder, "templates")
start_row = 6  # Data starts from row 6
yyyymm = datetime.now().strftime("%Y%m")
data_folder = os.path.join(base_folder, yyyymm)
output_folder = os.path.join(data_folder, "binarysheet")

# Build file map: xlsx → xlsb template
file_map = {}
for file in os.listdir(data_folder):
    if file.endswith(".xlsx"):
        match = re.match(r"(.+)_\d{6}\.xlsx", file)
        if match:
            base_name = match.group(1)
            template_name = f"{base_name}_template.xlsb"
            template_path = os.path.join(template_folder, template_name)
            if os.path.exists(template_path):
                file_map[os.path.join(data_folder, file)] = template_path
            else:
                print(f"⚠️ Template not found for {file} → expected: {template_name}")

# Process each xlsx file separatedly (new excel session each time)
for xlsx_path, xlsb_path in file_map.items():
    print(f"\n🔄 Importing {os.path.basename(xlsx_path)} → {os.path.basename(xlsb_path)}")

    #1 Load xlsx data using openpyxl
    try:
        wb_xlsx = load_workbook(xlsx_path, data_only=True)
        ws_xlsx = wb_xlsx.active
        data = [list(row) for row in ws_xlsx.iter_rows(values_only=True)]
        wb_xlsx.close()
    except Exception as e:
        print(f"❌ Failed to read {xlsx_path}: {e}")
        continue
    
    # 2. Start fresh Excel COM session
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.DisplayAlerts = False
        try:
            excel.Visible = False
        except:
            pass
    except Exception as e:
        print(f"❌ Excel failed to start: {e}")
        continue

    #3 Paste data into xlsb template via Excel
    try:
        wb = excel.Workbooks.Open(Filename=xlsb_path)
        ws = wb.Sheets(1)

        # Clear old data below headers
        last_row = ws.UsedRange.Rows.Count
        if last_row >= start_row:
            ws.Range(ws.Cells(start_row, 1), ws.Cells(last_row, ws_xlsx.max_column)).Clear()

        # Paste new data starting from row 6
        for i, row in enumerate(data):
            for j, val in enumerate(row):
                ws.Cells(start_row + i, j + 1).Value = val

        def clean_filename(filename):
            return re.sub(r'[\\/*?:"<>|]', "_", filename)

        # Remove "_template" from filename
        output_filename = clean_filename(os.path.basename(xlsb_path).replace("_template", ""))
        output_path = os.path.join(output_folder, output_filename)
        
        print(f"About to save. Output path: {output_path}")
        print(f"Does path exist? {os.path.exists(os.path.dirname(output_path))}")
        print(f"Valid extension? {output_path.endswith('.xlsb')}")

        # Save As
        wb.SaveAs(Filename=output_path, FileFormat=50)  # 50 = xlExcel12 (xlsb)
        wb.Close(SaveChanges=False)
        print(f"✅ Imported and saved as {output_filename} in 'binarysheet'")

    except Exception as e:
        print(f"❌ Failed to write to {xlsb_path}: {e}")

    #4 Final cleanup
    try:
        excel.Quit()
        del ws, wb, excel
        gc.collect()
        print("✅ Excel application closed.")
    except Exception as e:
        print(f"⚠️ Failed to close Excel gracefully: {e}")
        print("⛔ Forcing Excel shutdown...")